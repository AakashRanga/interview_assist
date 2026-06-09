from typing import Optional, List
from fastapi import APIRouter, HTTPException, Header, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from app.database import SessionLocal
from app.models.user import User
from app.services.excel_service import ExcelService
import openpyxl
import io
import json
from app.models.candidate import Candidate
from app.models.candidate_application import CandidateApplication
from app.models.interview_schedule import InterviewSchedule
from app.models.job_role import JobRole
from app.models.panel import Panel
from app.services.notification_service import NotificationService
from app.models.notification import ActivityType, NotificationType
from app.workers.tasks import schedule_interview_task
from datetime import datetime, date, time, timedelta
from zoneinfo import ZoneInfo

router = APIRouter(prefix="/admin", tags=["admin"])


class AddPanelRequest(BaseModel):
    full_name: str
    email: str
    password: str


class PanelResponse(BaseModel):
    id: int
    full_name: str
    email: str
    role: str

    class Config:
        orm_mode = True


class AdminResponse(BaseModel):
    status: str
    message: str
    panel: Optional[PanelResponse] = None


def get_current_user_from_email(email: str, db) -> Optional[User]:
    """Helper function to get user from email (in production, use JWT tokens)"""
    return db.query(User).filter(User.email == email).first()


@router.post("/add-panel", response_model=AdminResponse)
def add_panel(data: AddPanelRequest, admin_email: str = Header(None)):
    """
    Add a new panel member (interview panelist).
    Only admins can add panel members.
    admin_email should be passed as a header with the admin's email.
    """
    db = SessionLocal()
    try:
        # Verify admin
        if not admin_email:
            raise HTTPException(status_code=401, detail="Admin email required in headers")

        admin_user = get_current_user_from_email(admin_email, db)
        if not admin_user:
            raise HTTPException(status_code=401, detail="Admin not found")

        if admin_user.role != "admin":
            raise HTTPException(status_code=403, detail="Only admins can add panel members")

        # Check if email already exists
        existing_user = db.query(User).filter(User.email == data.email).first()
        if existing_user:
            raise HTTPException(status_code=400, detail="Email already registered")

        # Create new panel member
        panel_user = User(
            full_name=data.full_name,
            email=data.email,
            password=data.password,
            role="panel"
        )
        db.add(panel_user)
        db.commit()
        db.refresh(panel_user)

        return {
            "status": "success",
            "message": "Panel member added successfully",
            "panel": PanelResponse(
                id=panel_user.id,
                full_name=panel_user.full_name,
                email=panel_user.email,
                role=panel_user.role
            )
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()


# ============ Interview Scheduling ============

class ScheduleInterviewRequest(BaseModel):
    candidate_id: int
    job_id: Optional[int] = None


class ScheduleInterviewResponse(BaseModel):
    status: str
    message: str
    schedule_id: Optional[int] = None


def find_available_slot(db, target_date: date = None):
    """
    Find the next available slot for interview.
    Returns (start_time, end_time) or None if no slots available.
    Uses Indian Standard Time (IST).
    """
    # Get current time in IST
    IST = ZoneInfo('Asia/Kolkata')
    now_ist = datetime.now(IST)
    current_time_ist = now_ist.time()
    today_ist = now_ist.date()

    # If no target_date provided, use today (IST)
    if target_date is None:
        target_date = today_ist

    # If target_date is in the past, return None
    if target_date < today_ist:
        return None

    # Define working hours (9 AM to 5 PM IST)
    slot_duration = timedelta(hours=1)
    work_start = time(9, 0)
    work_end = time(17, 0)

    # Get all schedules for the target date
    schedules = db.query(InterviewSchedule).filter(
        InterviewSchedule.date == target_date,
        InterviewSchedule.interview_status != "cancelled"
    ).all()

    occupied_slots = [(s.start_time, s.end_time) for s in schedules]

    # Find first available slot
    current_time = work_start

    # If scheduling for today, start from the next available slot after current time
    if target_date == today_ist:
        # Check if current time is within working hours
        if current_time_ist >= work_start and current_time_ist < work_end:
            hour_now = now_ist.hour
            # If minute > 0, we're past the start of current hour slot, use next slot
            if current_time_ist.minute > 0:
                next_hour = hour_now + 1
                # If next hour is still within work hours
                if next_hour < 17:
                    current_time = time(next_hour, 0)
                else:
                    return None  # No more slots today
            # If minute == 0, we're exactly at slot start, can use current slot

    while current_time and current_time < work_end:
        slot_end = (datetime.combine(target_date, current_time) + slot_duration).time()

        # Check if this slot is occupied
        is_occupied = False
        for occupied_start, occupied_end in occupied_slots:
            # Check overlap
            if current_time < occupied_end and slot_end > occupied_start:
                is_occupied = True
                break

        if not is_occupied:
            return current_time, slot_end

        current_time = slot_end

    return None


@router.post("/schedule-interview", response_model=ScheduleInterviewResponse)
def schedule_interview(data: ScheduleInterviewRequest):
    """
    Schedule an interview for a candidate:
    1. Validate candidate exists and has email
    2. Find available slot
    3. Create schedule record
    4. Queue Celery task to call n8n
    """
    db = SessionLocal()

    try:
        # Step 1: Validate candidate
        candidate = db.query(Candidate).filter(
            Candidate.id == data.candidate_id
        ).first()

        if not candidate:
            raise HTTPException(status_code=404, detail="Candidate not found")

        if not candidate.email:
            raise HTTPException(status_code=400, detail="Candidate has no email address")

        # Resolve job_id if not provided
        job_id = data.job_id
        application = None
        job_type = "Online"
        venue = None
        if not job_id:
            application = db.query(CandidateApplication).filter(
                CandidateApplication.candidate_id == data.candidate_id
            ).order_by(CandidateApplication.created_at.desc()).first()
            if application:
                job_id = application.role_id
                if application.role:
                    job_type = application.role.job_type or "Online"
                    venue = application.role.venue
        else:
            # Check job_type directly from job role if provided
            job_role = db.query(JobRole).filter(JobRole.id == job_id).first()
            if job_role:
                job_type = job_role.job_type or "Online"
                venue = job_role.venue

        is_offline = job_type and job_type.lower() == "offline"

        # Step 3: Find available slot (uses IST)
        IST = ZoneInfo('Asia/Kolkata')
        now_ist = datetime.now(IST)
        target_date = now_ist.date()
        slot = find_available_slot(db, target_date)

        if not slot:
            # Try tomorrow
            target_date = now_ist.date() + timedelta(days=1)
            slot = find_available_slot(db, target_date)

        if not slot:
            raise HTTPException(status_code=400, detail="No available slots today or tomorrow")

        start_time, end_time = slot

        # Step 4: Create schedule record
        schedule = InterviewSchedule(
            candidate_id=data.candidate_id,
            job_id=job_id,
            date=target_date,
            start_time=start_time,
            end_time=end_time,
            interview_status="pending"  # Will be updated by Celery
        )
        db.add(schedule)
        db.commit()
        db.refresh(schedule)

        # Update candidate application status to "Scheduled"
        if not application and job_id:
            application = db.query(CandidateApplication).filter(
                CandidateApplication.candidate_id == data.candidate_id,
                CandidateApplication.role_id == job_id
            ).order_by(CandidateApplication.created_at.desc()).first()

        if application:
            application.status = "Scheduled"
            db.add(application)
            db.commit()

        # Update candidate status to "Scheduled"
        candidate.status = "Scheduled"
        db.add(candidate)
        db.commit()

        if application:
            try:
                NotificationService.create_activity_and_notify(
                    db=db,
                    candidate_id=candidate.id,
                    user_id=candidate.user_id,
                    activity_type=ActivityType.INTERVIEW_SCHEDULED,
                    activity_title=f"Interview scheduled for {application.role.title if application.role else 'your role'}",
                    notification_title="Interview Scheduled",
                    notification_message=f"Your interview has been scheduled for {target_date} from {start_time} to {end_time}.",
                    reference_id=application.id,
                    notification_type=NotificationType.SUCCESS,
                    icon="calendar",
                    priority="high",
                    redirect_url=f"/applications/{application.id}"
                )
            except Exception as notif_error:
                print(f"Failed to create notification: {notif_error}")

        # Step 5: Queue Celery task only for online interviews
        if not is_offline:
            schedule_interview_task.delay(schedule.id, data.candidate_id)
        else:
            # For offline interviews, just mark as scheduled without gmeet link
            schedule.interview_status = "scheduled"
            db.add(schedule)
            db.commit()

        if is_offline:
            return {
                "status": "scheduled",
                "message": f"Interview scheduled for {target_date} {start_time}-{end_time} at venue: {venue or 'TBD'}",
                "schedule_id": schedule.id,
                "venue": venue,
                "mode": "offline"
            }

        return {
            "status": "queued",
            "message": f"Interview scheduled for {target_date} {start_time}-{end_time}. n8n will generate Google Meet link.",
            "schedule_id": schedule.id
        }

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()


# ============ Candidate Management ============

class CandidateAdminDetail(BaseModel):
    id: str
    name: str
    email: str
    phone: Optional[str] = ""
    role: str
    status: str
    interviewDate: Optional[str] = None
    interviewTime: Optional[str] = None
    mode: str
    panelId: Optional[str] = ""
    panelGroupId: Optional[str] = ""
    panelIds: List[int] = []  # Panel IDs from interview schedule (JSON)
    profileComplete: int
    avatar: str
    skills: List[str]
    experience: str
    education: str
    resumeUrl: Optional[str] = None
    meetLink: Optional[str] = None
    venue: Optional[str] = None
    appliedDate: str
    roleId: Optional[int] = None


class UpdateCandidateStatusRequest(BaseModel):
    status: str


class ReassignCandidatePanelRequest(BaseModel):
    panel_id: str
    panel_group_id: str


class RescheduleInterviewRequest(BaseModel):
    date: str
    time: str


@router.get("/candidates", response_model=List[CandidateAdminDetail])
def get_admin_candidates():
    db = SessionLocal()
    try:
        candidates = db.query(Candidate).all()
        result = []
        for c in candidates:
            # 1. Fetch latest application
            application = db.query(CandidateApplication).filter(
                CandidateApplication.candidate_id == c.id
            ).order_by(CandidateApplication.created_at.desc()).first()
            
            role_title = "No Applied Role"
            app_status = c.status
            applied_date_str = c.created_at.isoformat() if c.created_at else datetime.utcnow().isoformat()
            job_type = "Online"
            venue = None
            job_id = None
            
            if application:
                role_title = application.role.title if application.role else "Unknown Role"
                app_status = application.status
                applied_date_str = application.created_at.isoformat()
                job_type = application.role.job_type if application.role else "Online"
                venue = application.role.venue if application.role else None
                job_id = application.role_id

            # 2. Fetch latest interview schedule
            sched_query = db.query(InterviewSchedule).filter(
                InterviewSchedule.candidate_id == c.id
            )
            if job_id:
                sched_query = sched_query.filter(InterviewSchedule.job_id == job_id)
            schedule = sched_query.order_by(InterviewSchedule.created_at.desc()).first()

            interview_date = None
            interview_time = None
            meet_link = None
            schedule_panel_ids = []

            if schedule:
                interview_date = schedule.date.isoformat() if schedule.date else None
                if schedule.start_time and schedule.end_time:
                    start_str = schedule.start_time.strftime("%I:%M %p")
                    end_str = schedule.end_time.strftime("%I:%M %p")
                    interview_time = f"{start_str} - {end_str}"
                elif schedule.start_time:
                    interview_time = schedule.start_time.strftime("%I:%M %p")
                meet_link = schedule.gmeet_link
                # Parse panel_ids from JSON string (e.g., "[1,2,3,4,5]")
                if schedule.panel_id:
                    try:
                        schedule_panel_ids = json.loads(schedule.panel_id)
                    except (json.JSONDecodeError, TypeError):
                        schedule_panel_ids = []

            # 3. Calculate profileComplete
            profile_pct = 60
            if c.phone:
                profile_pct += 10
            if c.resume_path:
                profile_pct += 10
            if c.education_list:
                profile_pct += 10
            if c.experience_list:
                profile_pct += 10
            if profile_pct > 100:
                profile_pct = 100

            # 4. Format skills
            skills_list = [s.strip() for s in c.skills.split(",") if s.strip()] if c.skills else []

            # 5. Format experience summary
            exp_str = "No experience listed"
            if c.experience_list:
                latest_exp = c.experience_list[0]
                exp_str = f"{latest_exp.years_experience} years as {latest_exp.current_role} at {latest_exp.company}"

            # 6. Format education summary
            edu_str = "No education listed"
            if c.education_list:
                latest_edu = c.education_list[0]
                edu_str = f"{latest_edu.degree} from {latest_edu.university} (Class of {latest_edu.graduation_year})"

            # 7. Resume URL
            resume_url = None
            if c.resume_path:
                if c.resume_path.startswith("http"):
                    resume_url = c.resume_path
                else:
                    # Return relative path - frontend will construct full URL
                    resume_url = c.resume_path

            avatar_url = f"https://api.dicebear.com/7.x/avataaars/svg?seed={c.id}"

            result.append(
                CandidateAdminDetail(
                    id=str(c.id),
                    name=c.full_name,
                    email=c.email,
                    phone=c.phone or "",
                    role=role_title,
                    status=app_status,
                    interviewDate=interview_date,
                    interviewTime=interview_time,
                    mode=job_type,
                    panelId=c.panel_id or "",
                    panelGroupId=c.panel_group_id or "",
                    panelIds=schedule_panel_ids,
                    profileComplete=profile_pct,
                    avatar=avatar_url,
                    skills=skills_list,
                    experience=exp_str,
                    education=edu_str,
                    resumeUrl=resume_url,
                    meetLink=meet_link,
                    venue=venue,
                    appliedDate=applied_date_str,
                    roleId=job_id
                )
            )
        return result
    finally:
        db.close()


@router.get("/candidates/export-excel")
def export_candidates_excel(
    search: Optional[str] = None,
    status: Optional[str] = None,
    panelGroupId: Optional[str] = None
):
    """
    Export matching candidates and their application details directly to Excel.
    """
    db = SessionLocal()
    try:
        candidates = db.query(Candidate).all()
        filtered_candidates = []
        
        for c in candidates:
            # 1. Fetch latest application
            application = db.query(CandidateApplication).filter(
                CandidateApplication.candidate_id == c.id
            ).order_by(CandidateApplication.created_at.desc()).first()
            
            role_id = ""
            role_name = ""
            app_status = c.status
            
            if application:
                role_id = application.role_id
                role_name = application.role.title if application.role else "Unknown Role"
                app_status = application.status

            # Match frontend filtering
            if search and search.lower() not in c.full_name.lower():
                continue
            if status and status != "all" and app_status != status:
                continue
            if panelGroupId and panelGroupId != "all" and c.panel_group_id != panelGroupId:
                continue

            filtered_candidates.append({
                "id": c.id,
                "name": c.full_name,
                "email": c.email,
                "mobile": c.phone or "",
                "role_id": role_id,
                "role_name": role_name,
                "status": app_status
            })
            
        # Create a new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Candidates"
        
        # Headers
        headers = ["Candidate ID", "Name", "Email", "Mobile", "Role ID", "Role Name", "Status"]
        ws.append(headers)
        
        for fc in filtered_candidates:
            ws.append([
                fc["id"],
                fc["name"],
                fc["email"],
                fc["mobile"],
                fc["role_id"],
                fc["role_name"],
                fc["status"]
            ])
            
        # Style headers
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        for col_idx in range(1, 8):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            
        # Set column widths automatically
        for col in ws.columns:
            vals = [cell.value for cell in col]
            max_len = max(len(str(v or '')) for v in vals)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"candidates_export_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        headers_resp = {
            'Content-Disposition': f'attachment; filename="{filename}"'
        }
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers_resp
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()


@router.put("/candidates/{candidate_id}/status")
def update_candidate_status(candidate_id: int, data: UpdateCandidateStatusRequest):
    db = SessionLocal()
    try:
        candidate = db.query(Candidate).filter(Candidate.id == candidate_id).first()
        if not candidate:
            raise HTTPException(status_code=404, detail="Candidate not found")
        
        # Update candidate status
        candidate.status = data.status
        db.add(candidate)
        
        # Also update latest application status
        application = db.query(CandidateApplication).filter(
            CandidateApplication.candidate_id == candidate_id
        ).order_by(CandidateApplication.created_at.desc()).first()
        
        if application:
            application.status = data.status
            db.add(application)
            
            try:
                NotificationService.create_activity_and_notify(
                    db=db,
                    candidate_id=candidate.id,
                    user_id=candidate.user_id,
                    activity_type=ActivityType.PROFILE_UPDATED,
                    activity_title=f"Application status updated to {data.status}",
                    notification_title="Application Status Updated",
                    notification_message=f"Your application status for {application.role.title if application.role else 'your role'} has been updated to {data.status}.",
                    reference_id=application.id,
                    notification_type=NotificationType.INFO,
                    icon="info",
                    priority="medium",
                    redirect_url=f"/profile"
                )
            except Exception as notif_error:
                print(f"Failed to create notification: {notif_error}")
        
        if data.status == "Selected":
            # Check if already scheduled to prevent duplicates
            existing_schedules = db.query(InterviewSchedule).filter(
                InterviewSchedule.candidate_id == candidate_id,
                InterviewSchedule.interview_status != "cancelled"
            ).first()
            if existing_schedules:
                raise HTTPException(
                    status_code=400,
                    detail="Candidate has already been scheduled for interviews."
                )

            # Find next free date D starting today/tomorrow
            IST = ZoneInfo('Asia/Kolkata')
            now_ist = datetime.now(IST)
            limit_time = time(10, 30)
            if now_ist.time() > limit_time:
                target_date = now_ist.date() + timedelta(days=1)
            else:
                target_date = now_ist.date()

            start_t = time(11, 0)
            end_t = time(12, 0)

            while True:
                conflict = db.query(InterviewSchedule).filter(
                    InterviewSchedule.candidate_id == candidate_id,
                    InterviewSchedule.date == target_date,
                    InterviewSchedule.interview_status != "cancelled",
                    InterviewSchedule.start_time < end_t,
                    InterviewSchedule.end_time > start_t
                ).first()
                if not conflict:
                    break
                target_date += timedelta(days=1)

            # Fetch the 5 panels
            panels = db.query(Panel).filter(
                Panel.interview_type.in_(["HR", "Technical", "P1", "P2", "P3"])
            ).all()

            panel_by_type = {}
            for p in panels:
                if p.interview_type not in panel_by_type:
                    panel_by_type[p.interview_type] = p

            job_id = application.role_id if application else None
            job_type = "Online"
            venue = None
            if application and application.role:
                job_type = application.role.job_type or "Online"
                venue = application.role.venue
            is_offline = job_type and job_type.lower() == "offline"

            # Collect all panel IDs for the 5-panel interview
            panel_ids = []
            for panel_type in ["HR", "Technical", "P1", "P2", "P3"]:
                panel = panel_by_type.get(panel_type)
                if panel:
                    panel_ids.append(panel.id)

            # Create ONE schedule with all panel IDs as JSON
            import json
            panel_ids_json = json.dumps(panel_ids) if panel_ids else None

            schedule = InterviewSchedule(
                candidate_id=candidate_id,
                job_id=job_id,
                date=target_date,
                start_time=start_t,
                end_time=end_t,
                panel_id=panel_ids_json,
                interview_status="pending"
            )
            db.add(schedule)
            db.flush()

            # Schedule background task only for online interviews (single task)
            if not is_offline:
                schedule_interview_task.delay(schedule.id, candidate_id)
            else:
                # For offline interviews, mark as scheduled without gmeet link
                schedule.interview_status = "scheduled"
                db.add(schedule)
                db.commit()

            try:
                if is_offline:
                    notif_message = f"Your parallel 5-panel interviews have been scheduled for {target_date} from 11:00 AM to 12:00 PM at venue: {venue or 'TBD'}."
                else:
                    notif_message = f"Your parallel 5-panel interviews have been scheduled for {target_date} from 11:00 AM to 12:00 PM."
                NotificationService.create_activity_and_notify(
                    db=db,
                    candidate_id=candidate.id,
                    user_id=candidate.user_id,
                    activity_type=ActivityType.INTERVIEW_SCHEDULED,
                    activity_title=f"Parallel interviews scheduled for {application.role.title if application and application.role else 'your role'}",
                    notification_title="Interviews Scheduled",
                    notification_message=notif_message,
                    reference_id=application.id if application else None,
                    notification_type=NotificationType.SUCCESS,
                    icon="calendar",
                    priority="high",
                    redirect_url=f"/profile"
                )
            except Exception as notif_error:
                print(f"Failed to create scheduling notification: {notif_error}")

        db.commit()
        return {"status": "success", "message": f"Candidate status updated to {data.status}"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()


@router.put("/candidates/{candidate_id}/reassign")
def reassign_candidate_panel(candidate_id: int, data: ReassignCandidatePanelRequest):
    db = SessionLocal()
    try:
        candidate = db.query(Candidate).filter(Candidate.id == candidate_id).first()
        if not candidate:
            raise HTTPException(status_code=404, detail="Candidate not found")
        
        candidate.panel_id = data.panel_id
        candidate.panel_group_id = data.panel_group_id
        db.add(candidate)
        
        try:
            NotificationService.create_activity_and_notify(
                db=db,
                candidate_id=candidate.id,
                user_id=candidate.user_id,
                activity_type=ActivityType.PROFILE_UPDATED,
                activity_title="Interview Panel Assigned",
                notification_title="Panel Assigned",
                notification_message="An interview panel group and category have been assigned to you.",
                reference_id=None,
                notification_type=NotificationType.INFO,
                icon="users",
                priority="medium",
                redirect_url=f"/profile"
            )
        except Exception as notif_error:
            print(f"Failed to create notification: {notif_error}")
            
        db.commit()
        return {"status": "success", "message": "Candidate panel reassigned successfully"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()


@router.put("/candidates/{candidate_id}/reschedule")
def reschedule_interview(candidate_id: int, data: RescheduleInterviewRequest):
    db = SessionLocal()
    try:
        candidate = db.query(Candidate).filter(Candidate.id == candidate_id).first()
        if not candidate:
            raise HTTPException(status_code=404, detail="Candidate not found")
        
        try:
            target_date = datetime.strptime(data.date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format. Expected YYYY-MM-DD")
            
        try:
            target_time = datetime.strptime(data.time.split()[0], "%H:%M").time()
        except ValueError:
            try:
                target_time = datetime.strptime(data.time.split()[0], "%H:%M:%S").time()
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid time format. Expected HH:MM")

        schedule = db.query(InterviewSchedule).filter(
            InterviewSchedule.candidate_id == candidate_id,
            InterviewSchedule.interview_status != "cancelled"
        ).order_by(InterviewSchedule.created_at.desc()).first()

        job_id = None
        application = db.query(CandidateApplication).filter(
            CandidateApplication.candidate_id == candidate_id
        ).order_by(CandidateApplication.created_at.desc()).first()
        if application:
            job_id = application.role_id

        # Check job_type for offline interviews
        job_type = "Online"
        venue = None
        if application and application.role:
            job_type = application.role.job_type or "Online"
            venue = application.role.venue
        is_offline = job_type and job_type.lower() == "offline"

        slot_duration = timedelta(hours=1)
        end_time = (datetime.combine(target_date, target_time) + slot_duration).time()

        if schedule:
            schedule.date = target_date
            schedule.start_time = target_time
            schedule.end_time = end_time
            schedule.interview_status = "pending"
            db.add(schedule)
        else:
            schedule = InterviewSchedule(
                candidate_id=candidate_id,
                job_id=job_id,
                date=target_date,
                start_time=target_time,
                end_time=end_time,
                interview_status="pending"
            )
            db.add(schedule)
        
        db.commit()
        db.refresh(schedule)

        if application and application.status != "Scheduled":
            application.status = "Scheduled"
            db.add(application)
            db.commit()

        if candidate.status != "Scheduled":
            candidate.status = "Scheduled"
            db.add(candidate)
            db.commit()

        try:
            if is_offline:
                notif_message = f"Your interview has been rescheduled for {target_date} at {target_time.strftime('%I:%M %p')} at venue: {venue or 'TBD'}."
            else:
                notif_message = f"Your interview has been rescheduled for {target_date} at {target_time.strftime('%I:%M %p')}."
            NotificationService.create_activity_and_notify(
                db=db,
                candidate_id=candidate.id,
                user_id=candidate.user_id,
                activity_type=ActivityType.INTERVIEW_SCHEDULED,
                activity_title=f"Interview rescheduled for {application.role.title if application and application.role else 'your role'}",
                notification_title="Interview Rescheduled",
                notification_message=notif_message,
                reference_id=application.id if application else None,
                notification_type=NotificationType.SUCCESS,
                icon="calendar",
                priority="high",
                redirect_url=f"/profile"
            )
        except Exception as notif_error:
            print(f"Failed to create notification: {notif_error}")

        if not is_offline:
            schedule_interview_task.delay(schedule.id, candidate.id)
        else:
            # For offline interviews, mark as scheduled without gmeet link
            schedule.interview_status = "scheduled"
            db.add(schedule)
            db.commit()

        if is_offline:
            return {
                "status": "success",
                "message": f"Interview rescheduled to {target_date} {target_time} at venue: {venue or 'TBD'}.",
                "schedule_id": schedule.id,
                "venue": venue,
                "mode": "offline"
            }

        return {
            "status": "success",
            "message": f"Interview rescheduled to {target_date} {target_time}. n8n will generate Google Meet link.",
            "schedule_id": schedule.id
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()


# ============ Roles Management ============

class AdminJobRoleResponse(BaseModel):
    id: str
    title: str
    location: str
    experience: str
    totalVacancy: int
    jobType: str
    venue: Optional[str] = None
    level: Optional[str] = None
    description: Optional[str] = None
    createdAt: str
    isVisible: bool


# ============ Excel Import ============

class ImportPanelResponse(BaseModel):
    status: str
    message: str
    imported_count: int
    failed_count: int
    errors: List[dict] = []


@router.post("/import-panels", response_model=ImportPanelResponse)
async def import_panels(file: UploadFile = File(...)):
    """
    Import panel members from Excel file.

    Column A: full_name
    Column B: email
    Column C: password
    Column D: role
    Column E: hr_panelists_name
    Column F: hr_panelist_grade
    Column G: hr_panel_mobile
    Column H: tag_coordinator
    Column I: slots
    Column J: team_link
    Column K: interview_type
    """

    db = SessionLocal()
    imported_count = 0
    failed_count = 0
    errors = []

    try:
        content = await file.read()
        workbook = openpyxl.load_workbook(io.BytesIO(content))
        worksheet = workbook.active

        for row_idx, row in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True),
            start=2
        ):
            try:
                full_name = row[0]
                email = row[1]
                password = row[2]
                role = row[3] or "panel"

                hr_panelists_name = row[4]
                hr_panelist_grade = row[5]
                hr_panel_mobile = row[6]
                tag_coordinator = row[7]
                slots = row[8]
                team_link = row[9]
                interview_type = row[10]

                if not full_name or not email or not password:
                    errors.append({
                        "row": row_idx,
                        "error": "Missing required fields: full_name, email, or password"
                    })
                    failed_count += 1
                    continue

                existing_user = (
                    db.query(User)
                    .filter(User.email == str(email).strip())
                    .first()
                )

                if existing_user:
                    errors.append({
                        "row": row_idx,
                        "error": f"Email {email} already exists"
                    })
                    failed_count += 1
                    continue

                user = User(
                    full_name=str(full_name).strip(),
                    email=str(email).strip(),
                    password=str(password).strip(),
                    role=str(role).strip()
                )

                db.add(user)
                db.flush()

                panel = Panel(
                    hr_panelists_name=str(hr_panelists_name).strip() if hr_panelists_name else None,
                    hr_panelist_grade=str(hr_panelist_grade).strip() if hr_panelist_grade else None,
                    hr_panel_mobile=str(hr_panel_mobile).strip() if hr_panel_mobile else None,
                    tag_coordinator=str(tag_coordinator).strip() if tag_coordinator else None,
                    slots=str(slots).strip() if slots else None,
                    team_link=str(team_link).strip() if team_link else None,
                    interview_type=str(interview_type).strip() if interview_type else None,
                )

                db.add(panel)
                db.commit()

                imported_count += 1

            except Exception as row_error:
                db.rollback()

                errors.append({
                    "row": row_idx,
                    "error": str(row_error)
                })

                failed_count += 1

        return {
            "status": "success" if failed_count == 0 else "partial",
            "message": f"Import completed. {imported_count} records imported, {failed_count} failed.",
            "imported_count": imported_count,
            "failed_count": failed_count,
            "errors": errors
        }

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

    finally:
        db.close()

@router.get("/download-panel-template")
def download_panel_template():
    """
    Download Excel template for importing panel members
    """
    try:
        template_io = ExcelService.create_panel_import_template()
        
        return StreamingResponse(
            iter([template_io.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=panel_import_template.xlsx"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


class CreateJobRoleRequest(BaseModel):
    title: str
    location: str
    experience: str
    totalVacancy: int
    jobType: str
    venue: Optional[str] = None
    level: Optional[str] = None
    description: Optional[str] = None


class ToggleVisibilityRequest(BaseModel):
    is_visible: bool


@router.get("/roles", response_model=List[AdminJobRoleResponse])
def get_admin_roles():
    db = SessionLocal()
    try:
        roles = db.query(JobRole).order_by(JobRole.id.desc()).all()
        result = []
        for r in roles:
            created_str = r.created_at.strftime("%Y-%m-%d") if r.created_at else datetime.utcnow().strftime("%Y-%m-%d")
            result.append(
                AdminJobRoleResponse(
                    id=str(r.id),
                    title=r.title,
                    location=r.location,
                    experience=r.experience,
                    totalVacancy=r.total_vacancy,
                    jobType=r.job_type,
                    venue=r.venue,
                    level=r.level,
                    description=r.description,
                    createdAt=created_str,
                    isVisible=r.is_visible
                )
            )
        return result
    finally:
        db.close()


@router.post("/roles")
def create_job_role(data: CreateJobRoleRequest):
    db = SessionLocal()
    try:
        role = JobRole(
            title=data.title,
            location=data.location,
            experience=data.experience,
            total_vacancy=data.totalVacancy,
            job_type=data.jobType,
            venue=data.venue,
            level=data.level,
            description=data.description,
            is_visible=True
        )
        db.add(role)
        db.commit()
        db.refresh(role)
        return {
            "status": "success",
            "message": "Job role created successfully",
            "role_id": role.id
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()


@router.put("/roles/{role_id}/visibility")
def toggle_job_role_visibility(role_id: int, data: ToggleVisibilityRequest):
    db = SessionLocal()
    try:
        role = db.query(JobRole).filter(JobRole.id == role_id).first()
        if not role:
            raise HTTPException(status_code=404, detail="Job role not found")
        
        role.is_visible = data.is_visible
        db.add(role)
        db.commit()
        
        status_text = "visible" if data.is_visible else "hidden"
        return {
            "status": "success",
            "message": f"Job role visibility updated to {status_text}"
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()


@router.delete("/roles/{role_id}")
def delete_job_role(role_id: int):
    db = SessionLocal()
    try:
        role = db.query(JobRole).filter(JobRole.id == role_id).first()
        if not role:
            raise HTTPException(status_code=404, detail="Job role not found")
        
        # Soft delete: set visibility to False to avoid foreign key issues
        role.is_visible = False
        db.add(role)
        db.commit()
        return {
            "status": "success",
            "message": "Job role deactivated successfully"
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()
