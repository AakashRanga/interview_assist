import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO


class ExcelService:
    """Service for handling Excel file operations"""

    @staticmethod
    def create_panel_import_template() -> BytesIO:
        """
        Create an Excel template for importing panel members.
        
        Columns:
        A: full_name (required)
        B: email (required)
        C: password (required)
        D: role (optional, default: 'panel')
        E: hr_panelists_name
        F: hr_panelist_grade
        G: hr_panel_mobile
        H: tag_coordinator
        I: slots
        J: team_link
        K: interview_type
        
        Returns:
            BytesIO: Excel file as bytes
        """
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Panels"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Add headers
        headers = [
            "full_name*",
            "email*",
            "password*",
            "role",
            "hr_panelists_name",
            "hr_panelist_grade",
            "hr_panel_mobile",
            "tag_coordinator",
            "slots",
            "team_link",
            "interview_type"
        ]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Add sample data rows
        sample_rows = [
            [
                "Rajesh Kumar",
                "rajesh.kumar@example.com",
                "Password123!",
                "panel",
                "Rajesh Kumar",
                "Senior Manager",
                "+91-9876543210",
                "Coordinator A",
                "5-10",
                "https://teams.microsoft.com/l/meetup-join/...",
                "Technical Round"
            ],
            [
                "Priya Singh",
                "priya.singh@example.com",
                "SecurePass456#",
                "panel",
                "Priya Singh",
                "Manager",
                "+91-9876543211",
                "Coordinator B",
                "10-15",
                "https://teams.microsoft.com/l/meetup-join/...",
                "HR Round"
            ],
            [
                "Amit Patel",
                "amit.patel@example.com",
                "MyPassword789$",
                "panel",
                "Amit Patel",
                "Lead",
                "+91-9876543212",
                "Coordinator C",
                "15-20",
                "https://teams.microsoft.com/l/meetup-join/...",
                "Final Round"
            ]
        ]

        for row_idx, row_data in enumerate(sample_rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = border
                if col_idx in [2, 7, 10]:  # Email, mobile, team_link - left aligned
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Add instructions sheet
        instructions_sheet = workbook.create_sheet("Instructions")
        instructions = [
            ["PANEL IMPORT TEMPLATE - INSTRUCTIONS", ""],
            ["", ""],
            ["REQUIRED FIELDS (marked with *):", ""],
            ["1. full_name", "Full name of the panel member"],
            ["2. email", "Valid email address (must be unique)"],
            ["3. password", "Initial password (will be used for login)"],
            ["", ""],
            ["OPTIONAL FIELDS:", ""],
            ["4. role", "Default: 'panel'. Can be: panel, hr, technical"],
            ["5. hr_panelists_name", "Panelist name for HR round"],
            ["6. hr_panelist_grade", "Grade/Level of the panelist"],
            ["7. hr_panel_mobile", "Contact number of HR panelist"],
            ["8. tag_coordinator", "Coordinator assigned to this panel"],
            ["9. slots", "Available interview slots (e.g., 5-10, 10-15)"],
            ["10. team_link", "Microsoft Teams or meeting link"],
            ["11. interview_type", "Type of interview (Technical, HR, Final Round, etc.)"],
            ["", ""],
            ["IMPORTANT NOTES:", ""],
            ["- Do not modify the header row", ""],
            ["- Email addresses must be unique across the system", ""],
            ["- Passwords should be strong and at least 8 characters", ""],
            ["- Fill in optional fields only if you have the information", ""],
            ["- Empty cells in optional columns will be left blank", ""],
            ["- Maximum 1000 records per import", ""],
        ]

        for row_idx, row_data in enumerate(instructions, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                cell = instructions_sheet.cell(row=row_idx, column=col_idx)
                cell.value = value
                if row_idx == 1:
                    cell.font = Font(bold=True, size=12, color="FFFFFF")
                    cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                elif "REQUIRED" in str(value) or "OPTIONAL" in str(value) or "IMPORTANT" in str(value):
                    cell.font = Font(bold=True, size=11)
                    cell.fill = PatternFill(start_color="D9E8F5", end_color="D9E8F5", fill_type="solid")

        instructions_sheet.column_dimensions['A'].width = 25
        instructions_sheet.column_dimensions['B'].width = 40

        # Set column widths for main sheet
        column_widths = [20, 28, 20, 12, 20, 18, 20, 20, 15, 35, 20]
        for col_idx, width in enumerate(column_widths, start=1):
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        # Freeze header row
        worksheet.freeze_panes = "A2"

        # Save to BytesIO
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output
